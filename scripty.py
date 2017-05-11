# -*- coding: utf-8 -*-
"""
    Identificacao: Alerta de risco de incendio
    Autor: Ronie Jovanhol
           Thiago Tuler
    Utilidades: Dado os preditores de entrada, Esse algoritmo fornece um
                percentual de risco de incendio por meio de um conjunto 
                de regras gerados por uma arvore de decisao.
"""
import os
import openpyxl

tree = {
    'index': 0, 
    'right': {
        'index': 6, 
        'right': {
            'index': 0, 
            'right': {
                'index': 3, 
                'right': 0.0214471, # terminal node 37
                'value': 1130.84399, 
                'left': {
                    'index': 2, 
                    'right': 0.0408473, # terminal node 36
                    'value': 5.29967, 
                    'left': 0.0818293 # terminal node 35
                }
            }, 
            'value': 6.08065, 
            'left': {
                'index': 1, 
                'right': {
                    'index': 0, 
                    'right': 0.175359, # terminal node 34
                    'value': 5.89440, 
                    'left': 0.0575384 # terminal node 33
                }, 
                'value': 624.57501, 
                'left': {
                    'index': 1, 
                    'right': {
                        'index': 0, 
                        'right': 0.202195, # terminal node 32
                        'value': 4.63185, 
                        'left': 0.0978987 # terminal node 31
                    }, 
                    'value': 506.88501, 
                    'left': 0.0315917 # terminal node 30
                }
            }
        }, 
        'value': 55.50000, 
        'left': {
            'index': 3, 
            'right': {
                'index': 7, #'index': 1, (Modificacao 06) 
                'right': 0.0473021, # terminal node 29
                'value': 24.66975, #'value': 736.80499, (Modificacao 06)
                'left': {
                    'index': 5, 
                    'right': {
                        'index': 3, 
                        'right': {
                            'index': 1, #'index': 0, (Modificacao 07)
                            'right': 0.586747, # terminal node 28
                            'value': 1086.23499, #'value': 24.23654, (Modificacao 07)
                            'left': 0.0750805 # terminal node 27
                        }, 
                        'value': 1371.60449, 
                        'left': 0.125505 # terminal node 26
                    }, 
                    'value': 6, 
                    'left': 0.0639006 # terminal node 25
                }
            }, 
            'value': 1169.50452, 
            'left': {
                'index': 7, 
                'right': 0.0699455, # terminal node 24
                'value': 25.03165, 
                'left': 0.185039 # terminal node 23
            }
        }
    }, 
    'value': 3.63463, 
    'left': {
        'index': 5, 
        'right': {
            'index': 3, 
            'right': {
                'index': 7, 
                'right': {
                    'index': 0, 
                    'right': {
                        'index': 3, #'index': 0, (Modificacao 05)
                        'right': 0.571113, # terminal node 22
                        'value': 1382.34058, #'value': 2.50642, (Modificacao 05)
                        'left': 0.239003 # terminal node 21
                    }, 
                    'value': 2.35704, 
                    'left': {
                        'index': 3, #'index': 7, (Modificacao 04)
                        'right': 0.338749, # terminal node 20
                        'value': 1400.65942, #'value': 24.71705, (Modificacao 04)
                        'left': 0.0901171 # terminal node 19
                    }
                }, 
                'value': 24.68275, 
                'left': {
                    'index': 1, 
                    'right': {
                        'index': 0, 
                        'right': 0.52072, # terminal node 18
                        'value': 2.35704, 
                        'left': 0.0606344 # terminal node 17
                    }, 
                    'value': 854.57001, 
                    'left': {
                        'index': 1, 
                        'right': 0.748276, # terminal node 16
                        'value': 684.70502, 
                        'left': 0.281141 # terminal node 15
                    }
                }
            }, 
            'value': 1350.50757, 
            'left': {
                'index': 1, 
                'right': 0.250526, # terminal node 14
                'value': 859.72498, 
                'left': {
                    'index': 3, #'index': 7, (Modificacao 03)
                    'right': 0.150805, # terminal node 13
                    'value': 1236.05347, #'value': 24.75305, (Modificacao 03)
                    'left': 0.0372407 # terminal node 12
                }
            }
        }, 
        'value': 8, 
        'left': {
            'index': 3, 
            'right': {
                'index': 0,
                'right': 0.139411, # terminal node 11
                'value': 3.26368, 
                'left': {
                    'index': 3, #'index': 4, (Modificacao 01)
                    'right': {
                        'index': 6,
                        'right': 0.0499697, # terminal node 10
                        'value': 44.50000, 
                        'left': {
                            'index': 3, #'index': 7, (Modificacao 02)
                            'right': 0.131192, # terminal node 09
                            'value': 1236.40845, #'value': 24.79185, (Modificacao 02)
                            'left': 0.0168376 # terminal node 08
                        }
                    },
                    'value': 1138.03552, #'value': 54.5, (Modificacao 01)
                    'left': 0.109868 # terminal node 07
                }
            }, 
            'value': 1109.07446, 
            'left': {
                'index': 1, 
                'right': {
                    'index': 1, 
                    'right': {
                        'index': 1, 
                        'right': 0.175612, # terminal node 06
                        'value': 864.86499, 
                        'left': 0.434419 # terminal node 05
                    }, 
                    'value': 841.71002, 
                    'left': 0.110923 # terminal node 04
                }, 
                'value': 556.09497, 
                'left': {
                    'index': 3, 
                    'right': {
                        'index': 3, 
                        'right': 0.18456, # terminal node 03
                        'value': 1090.74597, 
                        'left': 0.330305 # terminal node 02
                    }, 
                    'value': 1034.39551, 
                    'left': 0.0835317 # terminal node 01
                }
            }
        }
    }
}

class FireAlert (object):
    #Inicia a variavel relativa ao caminho dos arquivos utilizados
    def __init__(self):
        self.__location__ = os.path.realpath(
            os.path.join(os.getcwd(), os.path.dirname(__file__)))
    
    #Converte o nome da classe de uso da terra para um coeficiente numérico
    def str_to_int(self, uso_terra):
        if uso_terra == "Agricultura":
            return 1
        elif uso_terra == "Areas urbanas":
            return 2
        elif uso_terra == "Curso d'agua":
            return 3
        elif uso_terra == "Floresta natural":
            return 4
        elif uso_terra == "Solo Exposto":
            return 5
        elif uso_terra == "Pastagem":
            return 6
        elif uso_terra == "Silvicultura":
            return 7
        elif uso_terra =="Manguezais":
            return 8
        elif uso_terra == "Areas alagadas":
            return 9
        elif uso_terra == "Restinga":
            return 10
    
    # Faz uma predição com uma avore de decisão
    def predict(self, node, row):
        if row[node['index']] <= node['value']:
            if isinstance(node['left'], dict):
                return self.predict(node['left'], row)
            else:
                return node['left']
        else:
            if isinstance(node['right'], dict):
                return self.predict(node['right'], row)
            else:
                return node['right']
    
    #Utiliza os dados da tabela pra localizar o coeficiente de incêndio
    def process(self):
        print "Abrindo o arquivo..."
        workbook1 = openpyxl.load_workbook(os.path.join(self.__location__, 'dados_entrada.xlsx'), read_only=True)
        workbook2 = openpyxl.Workbook(write_only=True)
        print "Lendo a planilha..."
        worksheet1 = workbook1.active
        worksheet2 = workbook2.create_sheet()
        for index, cells in enumerate(worksheet1.iter_rows()):
            if index != 0:
                row = [float(cells[2].value), float(cells[3].value), float(cells[4].value), float(cells[5].value), float(cells[6].value), FireAlert().str_to_int(cells[7].value), float(cells[8].value), float(cells[9].value)]
                score = FireAlert().predict(tree, row)
                print 'Calculando {0} de {1}\r'.format(index+1, worksheet1.max_row),
                worksheet2.append([float(cells[0].value), float(cells[1].value), float(cells[2].value), float(cells[3].value), float(cells[4].value), float(cells[5].value), float(cells[6].value), cells[7].value, int(cells[8].value), float(cells[9].value), float(cells[10].value), score])
            else:
                worksheet2.append([cells[0].value, cells[1].value, cells[2].value, cells[3].value, cells[4].value, cells[5].value, cells[6].value, cells[7].value, cells[8].value, cells[9].value, cells[10].value, cells[11].value])
        workbook2.save(os.path.join(self.__location__, 'dados_saida.xlsx'))

if __name__ == '__main__':
    FireAlert().process()
